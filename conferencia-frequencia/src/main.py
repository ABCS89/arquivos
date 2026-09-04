# -*- coding: utf-8 -*-
"""
main.py - Conferência de Frequência

Uso:
    python src/main.py
        -> Detecta e processa automaticamente todos os pares em input/secretaria e input/sistema

    python src/main.py <pdf_secretaria> <pdf_sistema> [--mes AAAA-MM]
        -> Processa um par específico

Gera em output/:
  - conferencia_<CODIGO - NOME DA SECRETARIA>.xlsx  (aba "Retificações" + aba "Resumo")
  - retificacoes_<CODIGO - NOME DA SECRETARIA>.md   (se houver retificação)
  - retificacoes_<CODIGO - NOME DA SECRETARIA>.txt  (se houver retificação)
"""
import sys
import os
import re
from pathlib import Path
from datetime import datetime

# Garante que imports locais funcionem
sys.path.insert(0, str(Path(__file__).resolve().parent))

from extract_secretaria import extrair as extrair_secretaria, extrair_cabecalho
from extract_sistema import extrair as extrair_sistema
from compare import comparar_por_dia

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Diretórios base garantidos independentemente de onde o script é chamado
BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"

COR_CABECALHO = "1F4E78"
COR_SEM_REGISTRO_SECRETARIA = "FFF2CC"   # secretaria não tinha nada -> precisa lançar
COR_SEM_REGISTRO_SISTEMA = "D9E1F2"      # sistema não tinha nada -> conferir/lançar lá
COR_TIPO_DIFERENTE = "F8CBAD"            # os dois têm algo, mas tipos diferentes

MESES_PT = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
            "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]


def sanitizar_nome_arquivo(texto):
    """Remove caracteres proibidos no sistema de arquivos do Windows."""
    for ch in r'\/:*?"<>|':
        texto = texto.replace(ch, "-")
    return " ".join(texto.split()).strip(" .-")


def _cor_da_linha(r):
    if r["tipo_secretaria"] == "SEM REGISTRO":
        return COR_SEM_REGISTRO_SECRETARIA
    if r["tipo_sistema"] == "sem registro em sistema":
        return COR_SEM_REGISTRO_SISTEMA
    return COR_TIPO_DIFERENTE


def _texto_data(r):
    if r["data_inicio"] == r["data_fim"]:
        return r["data_inicio"].strftime("%d/%m/%Y")
    return f'{r["data_inicio"].strftime("%d/%m/%Y")} a {r["data_fim"].strftime("%d/%m/%Y")}'


def _linha_texto(r):
    return (f'{r["matricula"]} - {r["nome"]} - {_texto_data(r)} - {r["dias"]} '
            f'dia{"s" if r["dias"] != 1 else ""} - {r["tipo_secretaria"]} --> {r["tipo_sistema"]}')


def gerar_excel(retificacoes, caminho_saida, total_sec, total_sis, mes_label, nome_orgao=""):
    wb = Workbook()

    ws = wb.active
    ws.title = "Retificações"
    colunas = ["Matrícula", "Nome", "Data Início", "Data Fim", "Dias",
               "Tipo Atual (a corrigir)", "Tipo Correto"]
    ws.append(colunas)
    for cel in ws[1]:
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        cel.alignment = Alignment(horizontal="center")

    for r in retificacoes:
        ws.append([
            r["matricula"], r["nome"],
            r["data_inicio"].strftime("%d/%m/%Y"), r["data_fim"].strftime("%d/%m/%Y"),
            r["dias"], r["tipo_secretaria"], r["tipo_sistema"],
        ])
        cor = _cor_da_linha(r)
        for cel in ws[ws.max_row]:
            cel.fill = PatternFill("solid", fgColor=cor)

    larguras = [12, 34, 14, 14, 8, 30, 30]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Resumo")
    titulo_resumo = f"Conferência de Frequência — {nome_orgao} ({mes_label})" if nome_orgao else f"Conferência de Frequência ({mes_label})"
    ws2.append([titulo_resumo, ""])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append(["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws2.append(["Órgão / Secretaria", nome_orgao or "Não identificado"])
    ws2.append(["Mês de referência", mes_label])
    ws2.append(["Registros extraídos (secretaria)", total_sec])
    ws2.append(["Registros extraídos (sistema)", total_sis])
    ws2.append(["Total de retificações", len(retificacoes)])
    ws2.append([])
    ws2.append(["Legenda de cores", ""])
    ws2["A9"].font = Font(bold=True)
    ws2.append(["Amarelo", "secretaria não tinha nada lançado nesse dia (lançar lá)"])
    ws2.append(["Azul", "sistema não tem nada nesse dia (conferir se é pra lançar lá)"])
    ws2.append(["Laranja", "os dois têm algo lançado, mas de tipo diferente"])
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 55

    os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)
    wb.save(caminho_saida)


def gerar_txt(retificacoes, caminho_saida):
    linhas = [_linha_texto(r) for r in retificacoes]
    os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)
    with open(caminho_saida, "w", encoding="utf-8") as f:
        f.write("\n".join(linhas) + "\n")


def gerar_markdown(retificacoes, caminho_saida, mes_label, total_sec, total_sis, nome_orgao=""):
    linhas = []
    titulo = f"# Retificações de Frequência — {nome_orgao} ({mes_label})" if nome_orgao else f"# Retificações de Frequência ({mes_label})"
    linhas.append(titulo)
    linhas.append("")
    linhas.append(f"- Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    if nome_orgao:
        linhas.append(f"- Órgão / Secretaria: {nome_orgao}")
    linhas.append(f"- Mês de referência: {mes_label}")
    linhas.append(f"- Registros extraídos (secretaria): {total_sec}")
    linhas.append(f"- Registros extraídos (sistema): {total_sis}")
    linhas.append(f"- **Total de retificações: {len(retificacoes)}**")
    linhas.append("")
    linhas.append("Favor retificar as seguintes frequências:")
    linhas.append("")
    for r in retificacoes:
        linhas.append(f"- {_linha_texto(r)}")
    linhas.append("")

    os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)
    with open(caminho_saida, "w", encoding="utf-8") as f:
        f.write("\n".join(linhas))


def processar_par(caminho_secretaria, caminho_sistema, ano_mes=None):
    """Processa um par de relatórios (secretaria + sistema) e gera os arquivos em output/."""
    caminho_secretaria = Path(caminho_secretaria)
    caminho_sistema = Path(caminho_sistema)

    print("\n" + "=" * 65)
    print(f"PROCESSANDO CONFERÊNCIA:")
    print(f"  Secretaria: {caminho_secretaria.name}")
    print(f"  Sistema:    {caminho_sistema.name}")
    print("=" * 65)

    # 1. Extrair cabeçalho da secretaria (código, nome e mês/ano)
    codigo_sec, nome_sec, mes_cabecalho = extrair_cabecalho(str(caminho_secretaria))
    if ano_mes is None and mes_cabecalho:
        ano_mes = mes_cabecalho

    if codigo_sec and nome_sec:
        nome_orgao = f"{codigo_sec} - {nome_sec}"
    elif nome_sec:
        nome_orgao = nome_sec
    else:
        # Tenta pegar prefixo do arquivo (ex: "116 - secretaria.pdf" -> "116")
        m_arq = re.match(r"^(\d{3})\b", caminho_secretaria.stem)
        nome_orgao = m_arq.group(1) if m_arq else caminho_secretaria.stem

    print(f"Identificação do Órgão: {nome_orgao}")

    # 2. Extrair registros da secretaria
    print("Lendo registros da secretaria...")
    regs_sec, nao_reconhecidas = extrair_secretaria(str(caminho_secretaria))
    print(f"  -> {len(regs_sec)} registro(s) extraído(s)")
    if nao_reconhecidas:
        print(f"  -> {len(nao_reconhecidas)} linha(s) ignorada(s)/ruído")

    # 3. Extrair registros do sistema
    print("Lendo registros do sistema...")
    regs_sis = extrair_sistema(str(caminho_sistema))
    print(f"  -> {len(regs_sis)} registro(s) extraído(s)")

    # 4. Comparar dia a dia
    retificacoes, (ano_ref, mes_ref) = comparar_por_dia(regs_sec, regs_sis, ano_mes=ano_mes)
    mes_label = f"{MESES_PT[mes_ref]}/{ano_ref}" if (ano_ref and mes_ref) else "mês não identificado"
    print(f"Mês de referência apurado: {mes_label}")
    print(f"Total de retificações encontradas: {len(retificacoes)}")

    # 5. Definir nomes de saída
    nome_saida = sanitizar_nome_arquivo(nome_orgao)
    caminho_xlsx = OUTPUT_DIR / f"conferencia_{nome_saida}.xlsx"
    gerar_excel(retificacoes, str(caminho_xlsx), len(regs_sec), len(regs_sis), mes_label, nome_orgao=nome_orgao)
    print(f"[OK] Excel salvo em: {caminho_xlsx.relative_to(BASE_DIR)}")

    if retificacoes:
        caminho_md = OUTPUT_DIR / f"retificacoes_{nome_saida}.md"
        gerar_markdown(retificacoes, str(caminho_md), mes_label, len(regs_sec), len(regs_sis), nome_orgao=nome_orgao)
        print(f"[OK] Markdown salvo em: {caminho_md.relative_to(BASE_DIR)}")

        caminho_txt = OUTPUT_DIR / f"retificacoes_{nome_saida}.txt"
        gerar_txt(retificacoes, str(caminho_txt))
        print(f"[OK] Texto puro salvo em: {caminho_txt.relative_to(BASE_DIR)}")
    else:
        print("[INFO] Nenhuma retificação necessária (100% de conformidade!).")

    return {
        "orgao": nome_orgao,
        "mes": mes_label,
        "registros_sec": len(regs_sec),
        "registros_sis": len(regs_sis),
        "retificacoes": len(retificacoes),
    }


def descobrir_pares():
    """Varre as pastas input/secretaria e input/sistema e pareia os arquivos correspondentes."""
    sec_dir = INPUT_DIR / "secretaria"
    sis_dir = INPUT_DIR / "sistema"

    if not sec_dir.exists() or not sis_dir.exists():
        return []

    pares = []
    for f_sec in sorted(sec_dir.glob("*.pdf")):
        # 1. Padrão: "COD - secretaria.pdf" <-> "COD - sistema.pdf"
        m = re.match(r"^(.+?)\s*-\s*secretaria\.pdf$", f_sec.name, re.IGNORECASE)
        if m:
            prefixo = m.group(1).strip()
            # Procura no sistema por "COD - sistema.pdf"
            f_sis = sis_dir / f"{prefixo} - sistema.pdf"
            if f_sis.exists():
                pares.append((f_sec, f_sis))
                continue

        # 2. Padrão: Mesmo nome de arquivo em ambas as pastas (ex: "2026-04.pdf")
        f_sis_mesmo_nome = sis_dir / f_sec.name
        if f_sis_mesmo_nome.exists():
            pares.append((f_sec, f_sis_mesmo_nome))
            continue

        # 3. Padrão por código numérico inicial (ex: 116...)
        m_num = re.match(r"^(\d+)", f_sec.stem)
        if m_num:
            cod = m_num.group(1)
            sis_candidatos = list(sis_dir.glob(f"{cod}*.pdf"))
            if sis_candidatos:
                pares.append((f_sec, sis_candidatos[0]))

    return pares


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]

    ano_mes = None
    if "--mes" in sys.argv:
        valor = sys.argv[sys.argv.index("--mes") + 1]
        ano_str, mes_str = valor.split("-")
        ano_mes = (int(ano_str), int(mes_str))

    if len(args) == 2:
        caminho_secretaria, caminho_sistema = args
        processar_par(caminho_secretaria, caminho_sistema, ano_mes=ano_mes)
    elif len(args) == 0:
        pares = descobrir_pares()
        if not pares:
            print("\n[AVISO] Nenhum par de arquivos encontrado em:")
            print(f"  - {INPUT_DIR / 'secretaria'}")
            print(f"  - {INPUT_DIR / 'sistema'}")
            print("\nUso manual:")
            print("  python src/main.py <pdf_secretaria> <pdf_sistema> [--mes AAAA-MM]\n")
            sys.exit(1)

        print("\n" + "#" * 65)
        print(f"  CONFERÊNCIA DE FREQUÊNCIA — PROCESSAMENTO EM LOTE")
        print(f"  Foram encontrados {len(pares)} pares de relatórios para conferência.")
        print("#" * 65)

        resultados = []
        for sec, sis in pares:
            res = processar_par(sec, sis, ano_mes=ano_mes)
            resultados.append(res)

        print("\n" + "#" * 65)
        print("RESUMO FINAL DO PROCESSAMENTO:")
        print("#" * 65)
        for r in resultados:
            print(f"  * {r['orgao']} ({r['mes']}): {r['retificacoes']} retificação(ões)")
        print("#" * 65 + "\n")
    else:
        print("Uso: python src/main.py <pdf_secretaria> <pdf_sistema> [--mes AAAA-MM]")
        print("  Ou rode apenas 'python src/main.py' para processar todos os pares da pasta input/")
        sys.exit(1)


if __name__ == "__main__":
    main()
