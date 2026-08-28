"""
Etapa 3: preencher o cesta_basica.xlsx (aba consolidada + abas de secretaria)
e depois separar cada aba de secretaria em um arquivo .xlsx individual.
"""

import shutil
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

from . import config


def _limpar_area(ws, linha_inicio: int, linha_fim: int, colunas: int = 3):
    """Apaga o conteúdo de uma faixa de linhas antes de escrever os dados novos,
    para não deixar 'lixo' de uma execução anterior (ex.: mês passado tinha mais
    funcionários do que este mês)."""
    for linha in range(linha_inicio, linha_fim + 1):
        for col in range(1, colunas + 1):
            celula = ws.cell(row=linha, column=col)
            if not isinstance(celula, MergedCell):
                celula.value = None


def preencher_planilha_consolidada(wb, registros):
    """Escreve TODOS os funcionários na aba consolidada (Planilha1), a partir da linha 2."""
    ws = wb[config.ABA_CONSOLIDADA]

    limite = max(ws.max_row, len(registros) + 1)
    _limpar_area(ws, 2, limite)

    for i, registro in enumerate(registros, start=2):
        ws.cell(row=i, column=1, value=registro["funcional"])
        ws.cell(row=i, column=2, value=registro["nome"])
        ws.cell(row=i, column=3, value=registro["codigo_secretaria"])


def preencher_abas_secretarias(wb, grupos: dict):
    """
    Em cada aba de secretaria (todas exceto a consolidada), escreve a lista de
    funcionários daquela secretaria a partir da linha 3 — preservando o e-mail
    (linha 1) e o cabeçalho (linha 2) que já existem na aba.

    Devolve a lista de abas que ficaram sem nenhum funcionário neste mês
    (não é erro — só um aviso para você conferir).
    """
    abas_secretarias = [nome for nome in wb.sheetnames if nome != config.ABA_CONSOLIDADA]
    abas_sem_funcionarios = []

    for nome_aba in abas_secretarias:
        ws = wb[nome_aba]
        registros_da_aba = grupos.get(nome_aba, [])
        if not registros_da_aba:
            abas_sem_funcionarios.append(nome_aba)

        limite = max(ws.max_row, config.LINHA_INICIO_DADOS + len(registros_da_aba))
        _limpar_area(ws, config.LINHA_INICIO_DADOS, limite)

        for i, registro in enumerate(registros_da_aba, start=config.LINHA_INICIO_DADOS):
            ws.cell(row=i, column=1, value=registro["funcional"])
            ws.cell(row=i, column=2, value=registro["nome"])
            ws.cell(row=i, column=3, value=registro["codigo_secretaria"])

    return abas_sem_funcionarios


def exportar_arquivos_por_secretaria(caminho_cesta_atualizada: Path, competencia: str):
    """
    A partir do cesta_basica.xlsx já atualizado, gera um arquivo .xlsx separado
    para cada secretaria — cada um contendo só a aba daquela secretaria,
    pronto para anexar no e-mail.
    """
    config.SAIDA_SECRETARIAS_DIR.mkdir(parents=True, exist_ok=True)
    competencia_arquivo = competencia.replace("/", "-")

    wb_referencia = openpyxl.load_workbook(caminho_cesta_atualizada)
    abas_secretarias = [n for n in wb_referencia.sheetnames if n != config.ABA_CONSOLIDADA]
    wb_referencia.close()

    arquivos_gerados = {}
    for nome_aba in abas_secretarias:
        caminho_temp = config.SAIDA_SECRETARIAS_DIR / f"_tmp_{nome_aba}.xlsx"
        shutil.copy(caminho_cesta_atualizada, caminho_temp)

        wb_individual = openpyxl.load_workbook(caminho_temp)
        for outra_aba in list(wb_individual.sheetnames):
            if outra_aba != nome_aba:
                del wb_individual[outra_aba]

        nome_arquivo_final = f"Cesta Basica - {nome_aba} - {competencia_arquivo}.xlsx"
        caminho_final = config.SAIDA_SECRETARIAS_DIR / nome_arquivo_final
        wb_individual.save(caminho_final)
        wb_individual.close()
        caminho_temp.unlink()

        arquivos_gerados[nome_aba] = caminho_final

    return arquivos_gerados


def gerar_lista_envio(wb, grupos: dict, arquivos_gerados: dict, competencia: str) -> str:
    """Monta um texto em Markdown com e-mail, quantidade de funcionários e nome do
    arquivo de cada secretaria — para você copiar e colar na hora de enviar os e-mails."""
    linhas = [f"# Lista de envio - Cesta Básica ({competencia})", ""]

    abas_secretarias = sorted(n for n in wb.sheetnames if n != config.ABA_CONSOLIDADA)
    for nome_aba in abas_secretarias:
        ws = wb[nome_aba]
        email = ws.cell(row=config.LINHA_EMAIL, column=1).value or "(sem e-mail cadastrado)"
        qtd_funcionarios = len(grupos.get(nome_aba, []))
        arquivo = arquivos_gerados.get(nome_aba)
        nome_arquivo = arquivo.name if arquivo else "-"

        linhas.append(f"## Secretaria {nome_aba}")
        linhas.append(f"- E-mail: {email}")
        linhas.append(f"- Funcionários: {qtd_funcionarios}")
        linhas.append(f"- Arquivo: {nome_arquivo}")
        linhas.append("")

    return "\n".join(linhas)
