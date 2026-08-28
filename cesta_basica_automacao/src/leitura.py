"""
Etapa 1: encontrar os arquivos de entrada e ler os dados da Relação Mensal.
"""

import re
import unicodedata
from pathlib import Path

import openpyxl

from . import config


def _normalizar(texto: str) -> str:
    """Remove acentos e deixa em minúsculas, para comparar nomes de arquivo sem erro por acentuação."""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.lower()


def localizar_arquivos_entrada():
    """
    Procura, dentro da pasta entrada/, o arquivo da Relação Mensal e o cesta_basica.xlsx.

    Não depende do nome exato (ex.: 'Relação Mensal - 082026.xlsx' muda todo mês) —
    basta que o nome do arquivo contenha 'relacao'/'relação' ou 'cesta'.
    """
    candidatos = [p for p in config.ENTRADA_DIR.glob("*.xlsx") if not p.name.startswith("~$")]
    if not candidatos:
        raise FileNotFoundError(
            f"Nenhum arquivo .xlsx encontrado em {config.ENTRADA_DIR}. "
            "Coloque lá a Relação Mensal e o cesta_basica.xlsx."
        )

    arquivo_relacao = None
    arquivo_cesta = None
    for caminho in candidatos:
        nome_normalizado = _normalizar(caminho.stem)
        if "cesta" in nome_normalizado:
            arquivo_cesta = caminho
        elif "relacao" in nome_normalizado:
            arquivo_relacao = caminho

    if arquivo_relacao is None:
        raise FileNotFoundError(
            "Não encontrei o arquivo da Relação Mensal em entrada/ "
            "(o nome do arquivo deve conter 'Relação')."
        )
    if arquivo_cesta is None:
        raise FileNotFoundError(
            "Não encontrei o cesta_basica.xlsx em entrada/ "
            "(o nome do arquivo deve conter 'cesta')."
        )

    return arquivo_relacao, arquivo_cesta


def extrair_competencia(nome_arquivo: str) -> str:
    """
    Tenta extrair mês/ano do nome do arquivo da Relação Mensal.
    Ex.: 'Relação Mensal - 082026.xlsx' -> '08/2026'.
    Se não encontrar o padrão MMAAAA, devolve 'sem-data'.
    """
    m = re.search(r"(\d{2})(\d{4})(?!.*\d)", nome_arquivo)
    if m:
        mes, ano = m.groups()
        return f"{mes}/{ano}"
    return "sem-data"


def ler_relacao_mensal(caminho: Path):
    """
    Lê a Relação Mensal e devolve uma lista de dicionários apenas com as
    3 colunas que interessam: funcional, nome e código da secretaria.

    As colunas são localizadas pelo NOME do cabeçalho (não pela posição),
    então a ordem das colunas no arquivo pode mudar sem quebrar o script.
    """
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active  # primeira aba do arquivo

    cabecalho = [celula.value for celula in ws[1]]
    try:
        idx_funcional = cabecalho.index(config.COLUNA_FUNCIONAL)
        idx_nome = cabecalho.index(config.COLUNA_NOME)
        idx_secretaria = cabecalho.index(config.COLUNA_SECRETARIA)
    except ValueError as erro:
        raise ValueError(
            "Não encontrei uma das colunas esperadas "
            f"({config.COLUNA_FUNCIONAL!r}, {config.COLUNA_NOME!r}, {config.COLUNA_SECRETARIA!r}) "
            f"no cabeçalho da Relação Mensal. Cabeçalho encontrado: {cabecalho}"
        ) from erro

    registros = []
    for linha in ws.iter_rows(min_row=2, values_only=True):
        if linha[idx_funcional] is None:
            continue  # ignora linhas em branco no fim da planilha
        codigo_secretaria = str(linha[idx_secretaria]).strip()
        registros.append(
            {
                "funcional": linha[idx_funcional],
                "nome": linha[idx_nome],
                "codigo_secretaria": codigo_secretaria,
            }
        )

    wb.close()
    return registros
