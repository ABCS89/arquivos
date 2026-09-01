"""
Automação da Cesta Básica.

O que este script faz, passo a passo:
  1. Procura, na pasta entrada/, a Relação Mensal e o cesta_basica.xlsx.
  2. Lê da Relação Mensal as colunas Nro Funcional, Nome e Cód. Secretaria.
  3. Copia esses dados para a aba "Planilha1" do cesta_basica.xlsx (todos juntos)
     e distribui cada funcionário na aba da sua secretaria (a partir da linha 3,
     mantendo o e-mail na linha 1 e o cabeçalho na linha 2).
  4. Salva esse cesta_basica.xlsx atualizado em saida/.
  5. Gera, em saida/secretarias/, um arquivo .xlsx separado por secretaria —
     pronto para anexar no e-mail de cada uma.
  6. Gera saida/lista_envio.md com o e-mail e a quantidade de funcionários de
     cada secretaria, para facilitar o envio manual dos e-mails.

Uso:
    python main.py
"""

import openpyxl

from src import config, exportacao, leitura, processamento


def main():
    print("1) Localizando arquivos de entrada...")
    caminho_relacao, caminho_cesta = leitura.localizar_arquivos_entrada()
    competencia = leitura.extrair_competencia(caminho_relacao.name)
    print(f"   Relação Mensal: {caminho_relacao.name}")
    print(f"   cesta_basica:   {caminho_cesta.name}")
    print(f"   Competência detectada: {competencia}")

    print("\n2) Lendo a Relação Mensal...")
    registros = leitura.ler_relacao_mensal(caminho_relacao)
    print(f"   {len(registros)} funcionários lidos.")

    print("\n3) Agrupando por secretaria...")
    grupos = processamento.agrupar_por_secretaria(registros)
    print(f"   {len(grupos)} secretarias com funcionários neste mês.")

    print("\n4) Preenchendo o cesta_basica.xlsx...")
    wb = openpyxl.load_workbook(caminho_cesta)
    exportacao.preencher_planilha_consolidada(wb, registros)
    abas_sem_funcionarios = exportacao.preencher_abas_secretarias(wb, grupos)

    codigos_sem_aba = sorted(set(grupos.keys()) - set(wb.sheetnames))
    if codigos_sem_aba:
        print(
            f"   ATENÇÃO: os códigos de secretaria {codigos_sem_aba} aparecem na "
            "Relação Mensal, mas NÃO têm aba correspondente no cesta_basica.xlsx. "
            "Esses funcionários não foram distribuídos — crie a aba antes de reenviar."
        )
    if abas_sem_funcionarios:
        print(f"   Aviso: as abas {abas_sem_funcionarios} ficaram sem nenhum funcionário este mês.")

    config.SAIDA_DIR.mkdir(parents=True, exist_ok=True)
    competencia_arquivo = competencia.replace("/", "-")
    caminho_cesta_atualizada = config.SAIDA_DIR / f"cesta_basica - {competencia_arquivo}.xlsx"
    wb.save(caminho_cesta_atualizada)
    wb.close()
    print(f"   Arquivo consolidado salvo em: {caminho_cesta_atualizada}")

    print("\n5) Gerando um arquivo separado por secretaria...")
    arquivos_gerados = exportacao.exportar_arquivos_por_secretaria(caminho_cesta_atualizada, competencia, grupos)
    print(f"   {len(arquivos_gerados)} arquivos gerados em: {config.SAIDA_SECRETARIAS_DIR}")
    if abas_sem_funcionarios:
        print(f"   (as abas {abas_sem_funcionarios} não geraram arquivo, por estarem vazias este mês)")

    print("\n6) Gerando lista de envio (lista_envio.md)...")
    wb_leitura = openpyxl.load_workbook(caminho_cesta_atualizada)
    texto_lista = exportacao.gerar_lista_envio(wb_leitura, grupos, arquivos_gerados, competencia)
    wb_leitura.close()
    caminho_lista = config.SAIDA_DIR / "lista_envio.md"
    caminho_lista.write_text(texto_lista, encoding="utf-8")
    print(f"   Lista de envio salva em: {caminho_lista}")

    print("\nConcluído!")


if __name__ == "__main__":
    main()
