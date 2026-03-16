import pdfplumber
import pandas as pd
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ---------------------------------------------------
# NORMALIZAR OCORRENCIA
# ---------------------------------------------------

def normalizar_ocorrencia(oc):

    oc = oc.upper().strip()

    MAPA = {

        "FALTAS CLT": "FALTA",
        "FALTA": "FALTA",

        "ABONO ELEITORAL": "ABONO ELEITORAL",
        "ABONO": "ABONO",

        "TRATAMENTO DE SAÚDE": "TRATAMENTO DE SAUDE",
        "TRATAMENTO DE SAUDE": "TRATAMENTO DE SAUDE",

        "AUXÍLIO DOENÇA": "AUXILIO DOENCA",
        "AUXILIO DOENCA": "AUXILIO DOENCA",

        "DOENÇA EM PESSOA DA FAMÍLIA": "PESSOA DA FAMILIA",

        "ÔNUS PARA OUTRO ÓRGÃO": "ONUS PARA OUTRO ORGAO",

        "PESSOA DA": "PESSOA DA FAMILIA",
        "RETORNO/PERÍCIA": "PERICIA",
        "RETORNO/PERICIA": "PERICIA",

    }

    return MAPA.get(oc, oc)


# ---------------------------------------------------
# EXTRAI TEXTO DO PDF
# ---------------------------------------------------

def extrair_texto_pdf(caminho):

    texto = []

    with pdfplumber.open(caminho) as pdf:
        for pagina in pdf.pages:

            t = pagina.extract_text()

            if t:
                texto.append(t)

    return "\n".join(texto)

# ---------------------------------------------------
# CORRIGIR QUEBRA DE LINHAS DO PDF
# ---------------------------------------------------

def corrigir_quebras(texto):

    linhas = texto.split("\n")
    novas = []

    i = 0

    while i < len(linhas):

        linha = linhas[i].strip()

        # junta nomes quebrados antes da data
        if (
            i + 1 < len(linhas)
            and re.match(r'[A-ZÁ-Ú\s]+$', linha)
            and re.match(r'[A-ZÁ-Ú\s]+\s+\d{2}/\d{2}/\d{4}', linhas[i+1])
        ):
            linha = linha + " " + linhas[i+1].strip()
            i += 1

                # ignora rodapé do processo
        if "Peça do processo/documento PMP" in linha:
            i += 1
            continue

        if "materializada por:" in linha:
            i += 1
            continue

        if "CPF:" in linha:
            i += 1
            continue

        # casos como "Tratamento De" + "Saúde"
        if linha.endswith(" DE") and i + 1 < len(linhas):

            linha = linha + " " + linhas[i+1].strip()
            i += 1

        # casos como "Férias" + "Regulamentares"
        elif linha.endswith("FÉRIAS") and i + 1 < len(linhas):

            linha = linha + " " + linhas[i+1].strip()
            i += 1

        # casos como "Doação De" + "Sangue"
        elif linha.endswith("DOAÇÃO DE") and i + 1 < len(linhas):

            linha = linha + " " + linhas[i+1].strip()
            i += 1

        novas.append(linha)

        i += 1

    return "\n".join(novas)



# ---------------------------------------------------
# FREQUENCIA SECRETARIA
# ---------------------------------------------------

def parse_frequencia(texto):

    dados = []
    linhas = texto.split("\n")

    funcional_atual = None
    nome_atual = None

    padrao_funcionario = re.compile(r'(\d{2}\.\d{3}-\d)\s+([A-ZÁ-Ú\s]+)', re.I)

    padrao_ocorrencia = re.compile(
        r'([A-ZÁ-Úa-zá-ú\s]+)\s+(\d{2}/\d{2}/\d{4})\s+([\d,]+)'
    )

    for linha in linhas:

        m_func = padrao_funcionario.search(linha)

        if m_func:

            funcional_atual = m_func.group(1)
            nome_atual = m_func.group(2).strip()

            # palavras de ocorrência que podem aparecer grudadas no nome
            OCORRENCIAS = [
                "ABONO",
                "ABONO ELEITORAL",
                "FALTA",
                "FALTAS",
                "FÉRIAS REGULAMENTARES",
                "FERIAS REGULAMENTARES",
                "TRATAMENTO DE SAÚDE",
                "TRATAMENTO DE SAUDE",
                "AUXÍLIO DOENÇA",
                "AUXILIO DOENCA",
                "DOAÇÃO DE SANGUE",
                "DOACAO DE SANGUE",
                "FREQUÊNCIA NORMAL",
                "FREQUENCIA NORMAL"
            ]

            nome_upper = nome_atual.upper()

            for oc in OCORRENCIAS:
                if nome_upper.endswith(oc):
                    nome_atual = nome_atual[: -len(oc)].strip()
                    break

            continue

        m_oc = padrao_ocorrencia.search(linha)

        if m_oc and funcional_atual:

            ocorrencia = normalizar_ocorrencia(m_oc.group(1))

            qtd = float(m_oc.group(3).replace(",", "."))

            data = m_oc.group(2)

            dados.append({
                "funcional": funcional_atual,
                "nome": nome_atual,
                "data": data,
                "ocorrencia": ocorrencia,
                "qtd_secretaria": qtd
            })

    return pd.DataFrame(dados)


# ---------------------------------------------------
# RELATORIO RH
# ---------------------------------------------------

def parse_rh(texto):

    dados = []

    linhas = texto.split("\n")

    buffer = ""

    for linha in linhas:

        linha = linha.strip()

        if not linha:
            continue

        # ignora setor
        if linha.startswith("DIVISÃO"):
            continue

        # junta linhas quebradas
        buffer += " " + linha

        if re.search(r'\d{2}/\d{2}/\d{4}', linha):

            padrao = re.search(
                r'(\d{2}\.\d{3}-\d)\s+'
                r'([A-ZÁ-Ú\s]+?)\s+'
                r'(\d{2}/\d{2}/\d{4})\s+'
                r'(\d{2}/\d{2}/\d{4})\s+'
                r'(\d+)\s+'
                r'(.+)',
                buffer
            )

            if padrao:

                funcional = padrao.group(1)

                nome = padrao.group(2).strip()

                data = padrao.group(3)

                qtd = int(padrao.group(5))

                ocorrencia = normalizar_ocorrencia(
                    padrao.group(6).strip()
                )

                dados.append({
                    "funcional": funcional,
                    "nome": nome,
                    "data": data,
                    "ocorrencia": ocorrencia,
                    "qtd_rh": qtd
                })

            buffer = ""

    return pd.DataFrame(dados)
# ---------------------------------------------------
# COMPARAÇÃO (SEM DATA)
# ---------------------------------------------------

def comparar(df_freq, df_rh):

    sec = df_freq[["funcional","nome","data","ocorrencia"]].copy()
    sec = sec.rename(columns={"ocorrencia": "ocorrencia_secretaria"})

    rh = df_rh[["funcional","nome","data","ocorrencia"]].copy()
    rh = rh.rename(columns={"ocorrencia": "ocorrencia_rh"})

    merged = pd.merge(
        sec,
        rh,
        on=["funcional","nome","data"],
        how="outer"
    )

    def definir_status(row):

        if pd.isna(row["ocorrencia_secretaria"]):
            return "SÓ RH"

        if pd.isna(row["ocorrencia_rh"]):
            return "SÓ SECRETARIA"

        if row["ocorrencia_secretaria"] == row["ocorrencia_rh"]:
            return "OK"

        return "OCORRENCIA DIFERENTE"

    merged["status"] = merged.apply(definir_status, axis=1)

    return merged

# ---------------------------------------------------
# COLORIR EXCEL
# ---------------------------------------------------

def colorir_excel(arquivo):

    wb = load_workbook(arquivo)
    ws = wb.active

    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2):

        status = row[4].value

        if status == "DIVERGENTE":

            for cell in row:
                cell.fill = vermelho

    wb.save(arquivo)

# ---------------------------------------------------
# EXECUÇÃO
# ---------------------------------------------------

def main():

    pasta = Path(".")

    freq_pdf = list(pasta.glob("*frequencia_secretaria*.pdf"))[0]

    rh_pdf = list(pasta.glob("*relatorio_rh*.pdf"))[0]

    texto_freq = corrigir_quebras(extrair_texto_pdf(freq_pdf))

    texto_rh = corrigir_quebras(extrair_texto_pdf(rh_pdf))

    df_freq = parse_frequencia(texto_freq)

    df_rh = parse_rh(texto_rh)

    print("Secretaria:", len(df_freq))
    print("RH:", len(df_rh))

    print("\nOcorrencias Secretaria:")
    print(df_freq["ocorrencia"].value_counts())

    print("\nOcorrencias RH:")
    print(df_rh["ocorrencia"].value_counts())
    
    resultado = comparar(df_freq, df_rh)

    print("\nPrimeiras linhas RH:")
    print(df_rh.head(10))

    print("\nPrimeiras linhas Secretaria:")
    print(df_freq.head(10))

    arquivo_saida = "comparacao_frequencia.xlsx"

    resultado.to_excel(arquivo_saida, index=False)

    colorir_excel(arquivo_saida)

    print("\nConferência finalizada!\n")

    print(resultado)

    


if __name__ == "__main__":
    main()

