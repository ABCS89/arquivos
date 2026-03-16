# Biblioteca para ler texto de arquivos PDF
import pdfplumber

# Biblioteca para manipulação de tabelas (DataFrames)
import pandas as pd

# Biblioteca de expressões regulares (busca de padrões em texto)
import re

# Biblioteca para manipulação de datas
from datetime import datetime

# Biblioteca para trabalhar com caminhos e arquivos no sistema
from pathlib import Path

# Biblioteca para abrir e modificar arquivos Excel
from openpyxl import load_workbook

# Biblioteca para aplicar cores nas células do Excel
from openpyxl.styles import PatternFill


# ---------------------------------------------------
# NORMALIZAR OCORRENCIA
# ---------------------------------------------------

# Função responsável por padronizar os nomes das ocorrências
# Exemplo: "AUXÍLIO DOENÇA" -> "AUXILIO DOENCA"
def normalizar_ocorrencia(oc):

    # Converte texto para maiúsculo e remove espaços extras
    oc = oc.upper().strip()

    # Dicionário de conversão para padronizar os nomes
    MAPA = {

        "FALTAS CLT": "FALTA",
        "FALTA": "FALTA",

        "ABONO ELEITORAL": "ABONO ELEITORAL",
        "ABONO": "ABONO",

        "TRATAMENTO DE SAÚDE": "TRATAMENTO DE SAUDE",
        "TRATAMENTO DE SAUDE": "TRATAMENTO DE SAUDE",

        "AUXÍLIO DOENÇA": "AUXILIO DOENCA",
        "AUXILIO DOENCA": "AUXILIO DOENCA",

        "DOENÇA EM PESSOA DA FAMÍLIA": "DOENÇA EM PESSOA DA FAMÍLIA",

        "Cedido sem ônus para cedente": "Cedido sem ônus para cedente",
        "Aguardando retorno/perícia auxílio doenç": "Aguardando retorno/perícia auxílio doenç",

        "Aguardando perícia sempem": "Aguardando perícia sempem",

        "GALA": "GALA",
        "NOJO": "NOJO",
        "Suspensão - aposentadoria por invalidez": "Suspensão - aposentadoria por invalidez",

    }

    # Retorna o valor padronizado se existir no dicionário
    # Se não existir, retorna o valor original
    return MAPA.get(oc, oc)


# ---------------------------------------------------
# EXTRAI TEXTO DO PDF
# ---------------------------------------------------

# Função que abre um PDF e extrai todo o texto dele
def extrair_texto_pdf(caminho):

    # Lista onde será armazenado o texto de cada página
    texto = []

    # Abre o PDF
    with pdfplumber.open(caminho) as pdf:

        # Percorre todas as páginas do PDF
        for pagina in pdf.pages:

            # Extrai o texto da página
            t = pagina.extract_text()

            # Se encontrou texto na página
            if t:

                # Adiciona o texto à lista
                texto.append(t)

    # Junta todo o texto das páginas com quebra de linha
    return "\n".join(texto)


# ---------------------------------------------------
# CORRIGIR QUEBRA DE LINHAS DO PDF
# ---------------------------------------------------

# Muitos PDFs quebram linhas no meio das informações
# Esta função tenta reconstruir essas linhas corretamente
def corrigir_quebras(texto):

    # Divide o texto em linhas
    linhas = texto.split("\n")

    # Lista onde será armazenado o texto corrigido
    novas = []

    # Índice para percorrer as linhas
    i = 0

    # Loop manual para permitir pular linhas
    while i < len(linhas):

        # Remove espaços da linha atual
        linha = linhas[i].strip()

        # ------------------------------------------------
        # Junta nomes quebrados antes da data
        # Exemplo:
        # JOÃO DA SILVA
        # 12/05/2024 FALTA
        # ------------------------------------------------

        if (
            i + 1 < len(linhas)  # garante que existe próxima linha
            and re.match(r'[A-ZÁ-Ú\s]+$', linha)  # linha com nome
            and re.match(r'[A-ZÁ-Ú\s]+\s+\d{2}/\d{2}/\d{4}', linhas[i+1]) # próxima linha com data
        ):
            # junta as duas linhas
            linha = linha + " " + linhas[i+1].strip()

            # pula a próxima linha
            i += 1

        # ------------------------------------------------
        # Remove rodapés do sistema
        # ------------------------------------------------

        if "Peça do processo/documento PMP" in linha:
            i += 1
            continue

        if "materializada por:" in linha:
            i += 1
            continue

        if "CPF:" in linha:
            i += 1
            continue

        # ------------------------------------------------
        # Junta linhas quebradas de ocorrências
        # ------------------------------------------------

        # Exemplo:
        # TRATAMENTO DE
        # SAUDE

        if linha.endswith(" DE") and i + 1 < len(linhas):

            linha = linha + " " + linhas[i+1].strip()
            i += 1

        # Exemplo:
        # FÉRIAS
        # REGULAMENTARES

        elif linha.endswith("FÉRIAS") and i + 1 < len(linhas):

            linha = linha + " " + linhas[i+1].strip()
            i += 1

        # Exemplo:
        # DOAÇÃO DE
        # SANGUE

        elif linha.endswith("DOAÇÃO DE") and i + 1 < len(linhas):

            linha = linha + " " + linhas[i+1].strip()
            i += 1

        # Adiciona a linha corrigida na lista
        novas.append(linha)

        # Avança para próxima linha
        i += 1

    # Junta todas as linhas novamente
    return "\n".join(novas)



# ---------------------------------------------------
# FREQUENCIA SECRETARIA
# ---------------------------------------------------

# Esta função interpreta o PDF da Secretaria
def parse_frequencia(texto):

    # Lista onde serão armazenados os registros
    dados = []

    # Divide o texto em linhas
    linhas = texto.split("\n")

    # Variáveis para guardar o funcionário atual
    funcional_atual = None
    nome_atual = None

    # Padrão regex para identificar funcionário
    # exemplo: 12.345-6 JOÃO DA SILVA
    padrao_funcionario = re.compile(r'(\d{2}\.\d{3}-\d)\s+([A-ZÁ-Ú\s]+)', re.I)

    # Padrão para ocorrência
    # exemplo: FALTA 12/05/2024 1,0
    padrao_ocorrencia = re.compile(
        r'([A-ZÁ-Úa-zá-ú\s]+)\s+(\d{2}/\d{2}/\d{4})\s+([\d,]+)'
    )

    # Percorre todas as linhas
    for linha in linhas:

        # Procura funcionário
        m_func = padrao_funcionario.search(linha)

        if m_func:

            # Guarda o número funcional
            funcional_atual = m_func.group(1)

            # Guarda o nome
            nome_atual = m_func.group(2).strip()

            # Lista de ocorrências que podem aparecer grudadas no nome
            OCORRENCIAS = [
                "ABONO",
                "ABONO ELEITORAL",
                "FALTA",
                "FALTAS",
                "FÉRIAS REGULAMENTARES",
                "FERIAS REGULAMENTARES",
                "TRATAMENTO DE SAÚDE",
                "TRATAMENTO DE SAUDE",
                "AUXÍLIO DOENÇA",
                "AUXILIO DOENCA",
                "DOAÇÃO DE SANGUE",
                "DOACAO DE SANGUE",
                "FREQUÊNCIA NORMAL",
                "FREQUENCIA NORMAL",
                "Aguardando retorno/perícia auxílio doenç",
                "Gestante / maternidade",
                "Suspensão - aposentadoria por invalidez",
                "Minutos perdidos",
                "Gala",
                "NOJO",
                "Licença maternidade prorrogação",
                "Convocação judicial",
                "Aguardando perícia sempem",
                "Doença em pessoa da família",
                "Afastamento sem vencimentos",
                "Cedido sem ônus para cedente"
            ]

            nome_upper = nome_atual.upper()

            # Remove ocorrência grudada no nome
            for oc in OCORRENCIAS:
                if nome_upper.endswith(oc):
                    nome_atual = nome_atual[: -len(oc)].strip()
                    break

            continue

        # Procura ocorrência na linha
        m_oc = padrao_ocorrencia.search(linha)

        if m_oc and funcional_atual:

            # Padroniza o nome da ocorrência
            ocorrencia = normalizar_ocorrencia(m_oc.group(1))

            # Converte quantidade
            qtd = float(m_oc.group(3).replace(",", "."))

            # Pega a data
            data = m_oc.group(2)

            # Adiciona registro
            dados.append({
                "funcional": funcional_atual,
                "nome": nome_atual,
                "data": data,
                "ocorrencia": ocorrencia,
                "qtd_secretaria": qtd
            })

    # Retorna tabela pandas
    return pd.DataFrame(dados)


# ---------------------------------------------------
# RELATORIO RH
# ---------------------------------------------------

# Função que interpreta o relatório do RH
def parse_rh(texto):

    dados = []

    # divide o texto em linhas
    linhas = texto.split("\n")

    # regex que captura toda a linha do relatório RH
    padrao = re.compile(
        r'(\d{2}\.\d{3}-\d)\s+'      # funcional
        r'([A-ZÁ-Ú\s]+?)\s+'         # nome
        r'(\d{2}/\d{2}/\d{4})\s+'    # data início
        r'(\d{2}/\d{2}/\d{4})\s+'    # data fim
        r'(\d+)\s+'                  # quantidade
        r'(.+)',                     # ocorrência
        re.I
    )

    for linha in linhas:

        # tenta encontrar o padrão na linha
        m = padrao.search(linha)

        if not m:
            continue

        funcional = m.group(1)

        nome = m.group(2).strip()

        # converte datas
        data_inicio = datetime.strptime(m.group(3), "%d/%m/%Y")
        data_fim = datetime.strptime(m.group(4), "%d/%m/%Y")

        qtd = int(m.group(5))

        ocorrencia = normalizar_ocorrencia(
            m.group(6).strip()
        )

        # gera uma linha para cada dia do período
        data_atual = data_inicio

        while data_atual <= data_fim:

            dados.append({
                "funcional": funcional,
                "nome": nome,
                "data": data_atual.strftime("%d/%m/%Y"),
                "ocorrencia": ocorrencia,
                "qtd_rh": qtd
            })

            data_atual += pd.Timedelta(days=1)

    return pd.DataFrame(dados)


# ---------------------------------------------------
# COMPARAÇÃO
# ---------------------------------------------------

def comparar(df_freq, df_rh):

    # copia dados da secretaria
    sec = df_freq[["funcional","nome","data","ocorrencia"]].copy()

    # renomeia coluna
    sec = sec.rename(columns={"ocorrencia": "ocorrencia_secretaria"})

    # copia dados do RH
    rh = df_rh[["funcional","nome","data","ocorrencia"]].copy()

    rh = rh.rename(columns={"ocorrencia": "ocorrencia_rh"})

    # faz junção das duas tabelas
    merged = pd.merge(
        sec,
        rh,
        on=["funcional","nome","data"],
        how="outer"
    )

    # função para definir status
    def definir_status(row):

        if pd.isna(row["ocorrencia_secretaria"]):
            return "SÓ RH"

        if pd.isna(row["ocorrencia_rh"]):
            return "SÓ SECRETARIA"

        if row["ocorrencia_secretaria"] == row["ocorrencia_rh"]:
            return "OK"

        return "OCORRENCIA DIFERENTE"

    merged["status"] = merged.apply(definir_status, axis=1)

    return merged


# ---------------------------------------------------
# COLORIR EXCEL
# ---------------------------------------------------

def colorir_excel(arquivo):

    # abre arquivo excel
    wb = load_workbook(arquivo)

    ws = wb.active

    # define cor vermelha
    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # percorre linhas do excel
    for row in ws.iter_rows(min_row=2):

        status = row[4].value

        # pinta divergências
        if status == "OK":

            for cell in row:
                cell.fill = vermelho

    wb.save(arquivo)


# ---------------------------------------------------
# EXECUÇÃO PRINCIPAL
# ---------------------------------------------------

def main():

    # pasta atual do script
    pasta = Path(".")

    # procura arquivo de frequência
    freq_pdf = list(pasta.glob("*frequencia_secretaria*.pdf"))[0]

    # procura relatório do RH
    rh_pdf = list(pasta.glob("*relatorio_rh*.pdf"))[0]

    # extrai texto
    texto_freq = corrigir_quebras(extrair_texto_pdf(freq_pdf))

    texto_rh = corrigir_quebras(extrair_texto_pdf(rh_pdf))

    # converte para tabelas
    df_freq = parse_frequencia(texto_freq)

    df_rh = parse_rh(texto_rh)

    # imprime quantidades
    print("Secretaria:", len(df_freq))
    print("RH:", len(df_rh))

    # mostra ocorrências
    print("\nOcorrencias Secretaria:")
    print(df_freq["ocorrencia"].value_counts())

    print("\nOcorrencias RH:")
    print(df_rh["ocorrencia"].value_counts())

    # compara dados
    resultado = comparar(df_freq, df_rh)

    # mostra primeiras linhas
    print("\nPrimeiras linhas RH:")
    print(df_rh.head(10))

    print("\nPrimeiras linhas Secretaria:")
    print(df_freq.head(10))

    # define arquivo de saída
    arquivo_saida = "comparacao_frequencia.xlsx"

    # salva excel
    resultado.to_excel(arquivo_saida, index=False)

    # colore divergências
    colorir_excel(arquivo_saida)

    print("\nConferência finalizada!\n")

    print(resultado)


# executa o programa
if __name__ == "__main__":
    main()