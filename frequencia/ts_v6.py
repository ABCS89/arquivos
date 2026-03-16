import pdfplumber
import pandas as pd
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ---------------------------------------------------
# NORMALIZAR OCORRENCIA
# ---------------------------------------------------

def normalizar_ocorrencia(oc):

    oc = oc.upper().strip()

    MAPA = {

        "FALTAS EFETIVOS": "FALTA",
        "FALTA EFETIVOS": "FALTA",
        "FALTA": "FALTA",

        "ABONO ELEITORAL": "ABONO ELEITORAL",
        "ABONO": "ABONO",

        "TRATAMENTO DE SAÚDE": "TRATAMENTO DE SAUDE",
        "TRATAMENTO DE SAUDE": "TRATAMENTO DE SAUDE",

        "AUXILIO DOENÇA": "AUXILIO DOENCA",
        "AUXILIO DOENCA": "AUXILIO DOENCA",

    }

    return MAPA.get(oc, oc)


# ---------------------------------------------------
# EXTRAI TEXTO DO PDF
# ---------------------------------------------------

def extrair_texto_pdf(caminho):

    texto = []

    with pdfplumber.open(caminho) as pdf:
        for pagina in pdf.pages:

            t = pagina.extract_text()

            if t:
                texto.append(t)

    return "\n".join(texto)

# ---------------------------------------------------
# CORRIGIR QUEBRA DE LINHAS DO PDF
# ---------------------------------------------------

def corrigir_quebras(texto):

    linhas = texto.split("\n")
    novas = []

    i = 0

    while i < len(linhas):

        linha = linhas[i].strip()

        # casos como "Tratamento De" + "Saúde"
        if linha.endswith(" DE") and i + 1 < len(linhas):

            linha = linha + " " + linhas[i+1].strip()
            i += 1

        # casos como "Férias" + "Regulamentares"
        elif linha.endswith("FÉRIAS") and i + 1 < len(linhas):

            linha = linha + " " + linhas[i+1].strip()
            i += 1

        # casos como "Doação De" + "Sangue"
        elif linha.endswith("DOAÇÃO DE") and i + 1 < len(linhas):

            linha = linha + " " + linhas[i+1].strip()
            i += 1

        novas.append(linha)

        i += 1

    return "\n".join(novas)



# ---------------------------------------------------
# FREQUENCIA SECRETARIA
# ---------------------------------------------------

def parse_frequencia(texto):

    dados = []
    linhas = texto.split("\n")

    funcional_atual = None
    nome_atual = None

    padrao_funcionario = re.compile(r'(\d{2}\.\d{3}-\d)\s+([A-ZÁ-Ú\s]+)', re.I)

    padrao_ocorrencia = re.compile(
        r'([A-ZÁ-Úa-zá-ú\s]+)\s+(\d{2}/\d{2}/\d{4})\s+([\d,]+)'
    )

    for linha in linhas:

        m_func = padrao_funcionario.search(linha)

        if m_func:

            funcional_atual = m_func.group(1)
            nome_atual = m_func.group(2).strip()

            # palavras de ocorrência que podem aparecer grudadas no nome
            OCORRENCIAS = [
                "ABONO",
                "ABONO ELEITORAL",
                "FALTA",
                "FALTAS",
                "FÉRIAS REGULAMENTARES",
                "FERIAS REGULAMENTARES",
                "TRATAMENTO DE SAÚDE",
                "TRATAMENTO DE SAUDE",
                "AUXÍLIO DOENÇA",
                "AUXILIO DOENCA",
                "DOAÇÃO DE SANGUE",
                "DOACAO DE SANGUE",
                "FREQUÊNCIA NORMAL",
                "FREQUENCIA NORMAL"
            ]

            nome_upper = nome_atual.upper()

            for oc in OCORRENCIAS:
                if nome_upper.endswith(oc):
                    nome_atual = nome_atual[: -len(oc)].strip()
                    break

            continue

        m_oc = padrao_ocorrencia.search(linha)

        if m_oc and funcional_atual:

            ocorrencia = normalizar_ocorrencia(m_oc.group(1))

            qtd = float(m_oc.group(3).replace(",", "."))

            dados.append({
                "funcional": funcional_atual,
                "nome": nome_atual,
                "ocorrencia": ocorrencia,
                "qtd_secretaria": qtd
            })

    return pd.DataFrame(dados)


# ---------------------------------------------------
# RELATORIO RH
# ---------------------------------------------------

def parse_rh(texto):

    dados = []
    linhas = texto.split("\n")

    padrao = re.compile(
        r'(\d{2}\.\d{3}-\d)\s+(.+?)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(.+)'
    )

    for linha in linhas:

        m = padrao.search(linha)

        if not m:
            continue

        funcional = m.group(1)

        nome = m.group(2).strip()

        qtd = int(m.group(5))

        ocorrencia = normalizar_ocorrencia(m.group(6))

        dados.append({
            "funcional": funcional,
            "nome": nome,
            "ocorrencia": ocorrencia,
            "qtd_rh": qtd
        })

    return pd.DataFrame(dados)


# ---------------------------------------------------
# COMPARAÇÃO (SEM DATA)
# ---------------------------------------------------

def comparar(df_freq, df_rh):

    resultados = []

    # ocorrencias secretaria
    sec = df_freq[["funcional","nome","ocorrencia"]].drop_duplicates()

    # ocorrencias RH
    rh = df_rh[["funcional","nome","ocorrencia"]].drop_duplicates()

    merged = pd.merge(
        rh,
        sec,
        on=["funcional","ocorrencia"],
        how="outer",
        suffixes=("_rh","_sec"),
        indicator=True
    )

    for _, row in merged.iterrows():

        if row["_merge"] == "both":
            status = "OK"
        else:
            status = "DIVERGENTE"

        nome = row["nome_rh"] if pd.notna(row["nome_rh"]) else row["nome_sec"]

        resultados.append({

            "funcional": row["funcional"],
            "nome": nome,
            "ocorrencia": row["ocorrencia"],
            "status": status
        })

    return pd.DataFrame(resultados)

# ---------------------------------------------------
# COLORIR EXCEL
# ---------------------------------------------------

def colorir_excel(arquivo):

    wb = load_workbook(arquivo)
    ws = wb.active

    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2):

        status = row[3].value

        if status == "DIVERGENTE":

            for cell in row:
                cell.fill = vermelho

    wb.save(arquivo)

# ---------------------------------------------------
# EXECUÇÃO
# ---------------------------------------------------

def main():

    pasta = Path(".")

    freq_pdf = list(pasta.glob("*frequencia_secretaria*.pdf"))[0]

    rh_pdf = list(pasta.glob("*relatorio_rh*.pdf"))[0]

    texto_freq = corrigir_quebras(extrair_texto_pdf(freq_pdf))

    texto_rh = corrigir_quebras(extrair_texto_pdf(rh_pdf))

    df_freq = parse_frequencia(texto_freq)

    df_rh = parse_rh(texto_rh)

    print("Linhas secretaria:", len(df_freq))
    print("Linhas RH:", len(df_rh))

    resultado = comparar(df_freq, df_rh)

    arquivo_saida = "comparacao_frequencia.xlsx"

    resultado.to_excel(arquivo_saida, index=False)

    colorir_excel(arquivo_saida)

    print("\nConferência finalizada!\n")

    print(resultado)


if __name__ == "__main__":
    main()