import pdfplumber
import pandas as pd
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ---------------------------------------------------
# NORMALIZAR OCORRENCIA
# ---------------------------------------------------

def normalizar_ocorrencia(oc):

    oc = oc.upper().strip()

    MAPA = {

        "FALTAS EFETIVOS": "FALTA",
        "FALTA EFETIVOS": "FALTA",
        "FALTA": "FALTA",

        "ABONO ELEITORAL": "ABONO ELEITORAL",
        "ABONO": "ABONO",

        "TRATAMENTO DE SAÚDE": "TRATAMENTO DE SAUDE",
        "TRATAMENTO DE SAUDE": "TRATAMENTO DE SAUDE",

        "AUXILIO DOENÇA": "AUXILIO DOENCA",
        "AUXILIO DOENCA": "AUXILIO DOENCA",

    }

    return MAPA.get(oc, oc)


# ---------------------------------------------------
# EXTRAI TEXTO DO PDF
# ---------------------------------------------------

def extrair_texto_pdf(caminho):

    texto = []

    with pdfplumber.open(caminho) as pdf:
        for pagina in pdf.pages:

            t = pagina.extract_text()

            if t:
                texto.append(t)

    return "\n".join(texto)


# ---------------------------------------------------
# FREQUENCIA SECRETARIA
# ---------------------------------------------------

def parse_frequencia(texto):

    dados = []
    linhas = texto.split("\n")

    funcional_atual = None
    nome_atual = None

    padrao_funcionario = re.compile(r'(\d{2}\.\d{3}-\d)\s+([A-ZÁ-Ú\s]+)', re.I)

    padrao_ocorrencia = re.compile(
        r'([A-ZÁ-Úa-zá-ú\s]+)\s+(\d{2}/\d{2}/\d{4})\s+([\d,]+)'
    )

    for linha in linhas:

        m_func = padrao_funcionario.search(linha)

        if m_func:

            funcional_atual = m_func.group(1)
            nome_atual = m_func.group(2).strip()

            continue

        m_oc = padrao_ocorrencia.search(linha)

        if m_oc and funcional_atual:

            ocorrencia = normalizar_ocorrencia(m_oc.group(1))

            qtd = float(m_oc.group(3).replace(",", "."))

            dados.append({
                "funcional": funcional_atual,
                "nome": nome_atual,
                "ocorrencia": ocorrencia,
                "qtd_secretaria": qtd
            })

    return pd.DataFrame(dados)


# ---------------------------------------------------
# RELATORIO RH
# ---------------------------------------------------

def parse_rh(texto):

    dados = []
    linhas = texto.split("\n")

    padrao = re.compile(
        r'(\d{2}\.\d{3}-\d)\s+(.+?)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(.+)'
    )

    for linha in linhas:

        m = padrao.search(linha)

        if not m:
            continue

        funcional = m.group(1)

        nome = m.group(2).strip()

        qtd = int(m.group(5))

        ocorrencia = normalizar_ocorrencia(m.group(6))

        dados.append({
            "funcional": funcional,
            "nome": nome,
            "ocorrencia": ocorrencia,
            "qtd_rh": qtd
        })

    return pd.DataFrame(dados)


# ---------------------------------------------------
# COMPARAÇÃO (SEM DATA)
# ---------------------------------------------------

def comparar(df_freq, df_rh):

    resultados = []

    # agrupa secretaria
    freq_group = df_freq.groupby(
        ["funcional","ocorrencia"]
    ).agg(
        dias_secretaria=("qtd_secretaria","sum")
    ).reset_index()

    # agrupa RH
    rh_group = df_rh.groupby(
        ["funcional","ocorrencia"]
    ).agg(
        dias_rh=("qtd_rh","sum")
    ).reset_index()

    # merge
    merged = pd.merge(
        rh_group,
        freq_group,
        on=["funcional","ocorrencia"],
        how="outer"
    )

    merged = merged.fillna(0)

    for _, row in merged.iterrows():

        status = "OK"

        if row["dias_rh"] != row["dias_secretaria"]:
            status = "DIVERGENTE"

        resultados.append({

            "funcional": row["funcional"],
            "ocorrencia": row["ocorrencia"],
            "dias_rh": row["dias_rh"],
            "dias_secretaria": row["dias_secretaria"],
            "status": status
        })

    return pd.DataFrame(resultados)


# ---------------------------------------------------
# COLORIR EXCEL
# ---------------------------------------------------

def colorir_excel(arquivo):

    wb = load_workbook(arquivo)
    ws = wb.active

    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2):

        status = row[4].value

        if status == "DIVERGENTE":

            row[2].fill = vermelho
            row[3].fill = vermelho
            row[4].fill = vermelho

    wb.save(arquivo)


# ---------------------------------------------------
# EXECUÇÃO
# ---------------------------------------------------

def main():

    pasta = Path(".")

    freq_pdf = list(pasta.glob("*frequencia_secretaria*.pdf"))[0]

    rh_pdf = list(pasta.glob("*relatorio_rh*.pdf"))[0]

    texto_freq = extrair_texto_pdf(freq_pdf)

    texto_rh = extrair_texto_pdf(rh_pdf)

    df_freq = parse_frequencia(texto_freq)

    df_rh = parse_rh(texto_rh)

    print("Linhas secretaria:", len(df_freq))
    print("Linhas RH:", len(df_rh))

    resultado = comparar(df_freq, df_rh)

    arquivo_saida = "comparacao_frequencia.xlsx"

    resultado.to_excel(arquivo_saida, index=False)

    colorir_excel(arquivo_saida)

    print("\nConferência finalizada!\n")

    print(resultado)


if __name__ == "__main__":
    main()