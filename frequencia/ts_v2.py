import pdfplumber
import pandas as pd
import re
from datetime import datetime
from dateutil.relativedelta import relativedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


MES_REFERENCIA = 2
ANO_REFERENCIA = 2026


# ---------------------------------------------------
# EXTRAI TEXTO DO PDF
# ---------------------------------------------------

def extrair_texto_pdf(caminho):

    texto = []

    with pdfplumber.open(caminho) as pdf:
        for pagina in pdf.pages:

            t = pagina.extract_text()

            if t:
                texto.append(t)

    return "\n".join(texto)


# ---------------------------------------------------
# FREQUENCIA SECRETARIA
# ---------------------------------------------------

def parse_frequencia(texto):

    dados = []
    linhas = texto.split("\n")

    padrao = re.compile(
        r'(\d{2}\.\d{3}-\d)\s+(.+?)\s+(Férias regulamentares|Frequência normal)\s*(\d{2}/\d{2}/\d{4})?\s*([\d,]+)?',
        re.I
    )

    for linha in linhas:

        m = padrao.search(linha)

        if not m:
            continue

        funcional = m.group(1)
        nome = m.group(2).strip()
        ocorrencia = m.group(3).strip()

        data = m.group(4)
        qtd = m.group(5)

        if ocorrencia.lower() == "frequência normal":
            continue

        if data:
            data = datetime.strptime(data, "%d/%m/%Y")

        if qtd:
            qtd = float(qtd.replace(",", "."))

        dados.append({
            "funcional": funcional,
            "nome": nome,
            "ocorrencia": ocorrencia.upper(),
            "data": data,
            "qtd_secretaria": qtd
        })

    return pd.DataFrame(dados)


# ---------------------------------------------------
# RELATORIO RH
# ---------------------------------------------------

def parse_rh(texto):

    dados = []
    linhas = texto.split("\n")

    padrao = re.compile(
        r'(\d{2}\.\d{3}-\d)\s+(.+?)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(Férias Regulamentares)',
        re.I
    )

    for linha in linhas:

        m = padrao.search(linha)

        if not m:
            continue

        funcional = m.group(1)
        nome = m.group(2).strip()

        data_inicio = datetime.strptime(m.group(3), "%d/%m/%Y")
        data_fim = datetime.strptime(m.group(4), "%d/%m/%Y")

        qtd = int(m.group(5))
        ocorrencia = m.group(6).upper()

        dados.append({
            "funcional": funcional,
            "nome": nome,
            "ocorrencia": ocorrencia,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "qtd_rh": qtd
        })

    return pd.DataFrame(dados)


# ---------------------------------------------------
# CALCULAR DIAS DENTRO DO MES
# ---------------------------------------------------

def dias_no_mes(data_inicio, data_fim):

    inicio_mes = datetime(ANO_REFERENCIA, MES_REFERENCIA, 1)

    fim_mes = inicio_mes + relativedelta(months=1) - relativedelta(days=1)

    inicio = max(data_inicio, inicio_mes)
    fim = min(data_fim, fim_mes)

    if inicio > fim:
        return 0

    return (fim - inicio).days + 1


# ---------------------------------------------------
# COMPARAÇÃO
# ---------------------------------------------------

def comparar(df_freq, df_rh):

    resultados = []

    for _, rh in df_rh.iterrows():

        funcional = rh["funcional"]

        dias_mes = dias_no_mes(
            rh["data_inicio"],
            rh["data_fim"]
        )

        freq = df_freq[df_freq["funcional"] == funcional]

        if not freq.empty:

            qtd_secretaria = freq.iloc[0]["qtd_secretaria"]
            ocorrencia_secretaria = freq.iloc[0]["ocorrencia"]

        else:

            qtd_secretaria = 0
            ocorrencia_secretaria = "NÃO ENCONTRADO"

        status_dias = "OK"
        status_ocorrencia = "OK"

        if dias_mes != qtd_secretaria:
            status_dias = "DIVERGENTE"

        if rh["ocorrencia"] != ocorrencia_secretaria:
            status_ocorrencia = "DIVERGENTE"

        resultados.append({
            "funcional": funcional,
            "nome": rh["nome"],
            "ocorrencia_rh": rh["ocorrencia"],
            "ocorrencia_secretaria": ocorrencia_secretaria,
            "dias_rh_no_mes": dias_mes,
            "dias_secretaria": qtd_secretaria,
            "status_dias": status_dias,
            "status_ocorrencia": status_ocorrencia
        })

    return pd.DataFrame(resultados)


# ---------------------------------------------------
# COLORIR EXCEL
# ---------------------------------------------------

def colorir_excel(arquivo):

    wb = load_workbook(arquivo)
    ws = wb.active

    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2):

        status_dias = row[6].value
        status_ocorrencia = row[7].value

        if status_dias == "DIVERGENTE":
            row[4].fill = vermelho
            row[5].fill = vermelho

        if status_ocorrencia == "DIVERGENTE":
            row[2].fill = vermelho
            row[3].fill = vermelho

    wb.save(arquivo)


# ---------------------------------------------------
# EXECUÇÃO
# ---------------------------------------------------

def main():

    pasta = Path(".")

    freq_pdf = list(pasta.glob("*frequencia_secretaria*.pdf"))[0]
    rh_pdf = list(pasta.glob("*relatorio_rh*.pdf"))[0]

    texto_freq = extrair_texto_pdf(freq_pdf)
    texto_rh = extrair_texto_pdf(rh_pdf)

    df_freq = parse_frequencia(texto_freq)
    df_rh = parse_rh(texto_rh)

    resultado = comparar(df_freq, df_rh)

    arquivo_saida = "comparacao_frequencia.xlsx"

    resultado.to_excel(arquivo_saida, index=False)

    colorir_excel(arquivo_saida)

    print("\nConferência finalizada!\n")
    print(resultado)


if __name__ == "__main__":
    main()