import pdfplumber
import pandas as pd
import re
from datetime import datetime
from dateutil.relativedelta import relativedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


MES_REFERENCIA = 2
ANO_REFERENCIA = 2026


# ---------------------------------------------------
# EXTRAI TEXTO DO PDF
# ---------------------------------------------------

def extrair_texto_pdf(caminho):

    texto = []

    with pdfplumber.open(caminho) as pdf:
        for pagina in pdf.pages:

            t = pagina.extract_text()

            if t:
                texto.append(t)

    return "\n".join(texto)


# ---------------------------------------------------
# FREQUENCIA SECRETARIA
# ---------------------------------------------------

def parse_frequencia(texto):

    dados = []
    linhas = texto.split("\n")

    funcional_atual = None
    nome_atual = None

    padrao_funcionario = re.compile(r'(\d{2}\.\d{3}-\d)\s+([A-ZÁ-Ú\s]+)', re.I)

    padrao_ocorrencia = re.compile(
        r'([A-ZÁ-Úa-zá-ú\s]+)\s+(\d{2}/\d{2}/\d{4})\s+([\d,]+)'
    )

    for linha in linhas:

        m_func = padrao_funcionario.search(linha)

        if m_func:

            funcional_atual = m_func.group(1)
            nome_atual = m_func.group(2).strip()

            continue

        m_oc = padrao_ocorrencia.search(linha)

        if m_oc and funcional_atual:

            ocorrencia = m_oc.group(1).strip().upper()

            data = datetime.strptime(m_oc.group(2), "%d/%m/%Y")

            qtd = float(m_oc.group(3).replace(",", "."))

            dados.append({
                "funcional": funcional_atual,
                "nome": nome_atual,
                "ocorrencia": ocorrencia,
                "data": data,
                "qtd_secretaria": qtd
            })

    return pd.DataFrame(dados)


# ---------------------------------------------------
# RELATORIO RH
# ---------------------------------------------------

def parse_rh(texto):

    dados = []
    linhas = texto.split("\n")

    padrao = re.compile(
        r'(\d{2}\.\d{3}-\d)\s+(.+?)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(.+)'
    )

    for linha in linhas:

        m = padrao.search(linha)

        if not m:
            continue

        funcional = m.group(1)

        nome = m.group(2).strip()

        data_inicio = datetime.strptime(m.group(3), "%d/%m/%Y")

        data_fim = datetime.strptime(m.group(4), "%d/%m/%Y")

        qtd = int(m.group(5))

        ocorrencia = m.group(6).strip().upper()

        dados.append({
            "funcional": funcional,
            "nome": nome,
            "ocorrencia": ocorrencia,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "qtd_rh": qtd
        })

    return pd.DataFrame(dados)


# ---------------------------------------------------
# CALCULAR DIAS DENTRO DO MES
# ---------------------------------------------------

def dias_no_mes(data_inicio, data_fim):

    inicio_mes = datetime(ANO_REFERENCIA, MES_REFERENCIA, 1)

    fim_mes = inicio_mes + relativedelta(months=1) - relativedelta(days=1)

    inicio = max(data_inicio, inicio_mes)
    fim = min(data_fim, fim_mes)

    if inicio > fim:
        return 0

    return (fim - inicio).days + 1


# ---------------------------------------------------
# COMPARAÇÃO
# ---------------------------------------------------

def comparar(df_freq, df_rh):

    resultados = []

    # secretaria
    freq_group = df_freq.groupby(
        ["funcional","data","ocorrencia"]
    ).agg(
        dias_secretaria=("qtd_secretaria","sum")
    ).reset_index()

    # RH
    df_rh["data"] = df_rh["data_inicio"]

    rh_group = df_rh.groupby(
        ["funcional","data","ocorrencia"]
    ).agg(
        dias_rh=("qtd_rh","sum")
    ).reset_index()

    # merge
    merged = pd.merge(
        rh_group,
        freq_group,
        on=["funcional","data","ocorrencia"],
        how="outer"
    )

    merged = merged.fillna(0)

    for _, row in merged.iterrows():

        status_dias = "OK"
        status_ocorrencia = "OK"

        if row["dias_rh"] != row["dias_secretaria"]:
            status_dias = "DIVERGENTE"

        resultados.append({

            "funcional": row["funcional"],
            "data": row["data"],
            "ocorrencia": row["ocorrencia"],
            "dias_rh": row["dias_rh"],
            "dias_secretaria": row["dias_secretaria"],
            "status_dias": status_dias
        })

    return pd.DataFrame(resultados)

# ---------------------------------------------------
# CALCULAR DIAS DENTRO DO MES
# ---------------------------------------------------

def normalizar_ocorrencia(oc):

    oc = oc.upper().strip()

    MAPA = {

        "FALTAS EFETIVOS": "FALTA",
        "FALTA EFETIVOS": "FALTA",
        "FALTA": "FALTA",

        "ABONO ELEITORAL": "ABONO ELEITORAL",
        "ABONO": "ABONO",

        "TRATAMENTO DE SAÚDE": "TRATAMENTO DE SAUDE",
        "TRATAMENTO DE SAUDE": "TRATAMENTO DE SAUDE",

        "AUXILIO DOENÇA": "AUXILIO DOENCA",
        "AUXILIO DOENCA": "AUXILIO DOENCA",

    }

    return MAPA.get(oc, oc)

# ---------------------------------------------------
# COLORIR EXCEL
# ---------------------------------------------------

def colorir_excel(arquivo):

    wb = load_workbook(arquivo)
    ws = wb.active

    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2):

        status_dias = row[5].value

        if status_dias == "DIVERGENTE":

            row[3].fill = vermelho
            row[4].fill = vermelho
            row[5].fill = vermelho

    wb.save(arquivo)

# ---------------------------------------------------
# EXECUÇÃO
# ---------------------------------------------------

def main():

    pasta = Path(".")

    freq_pdf = list(pasta.glob("*frequencia_secretaria*.pdf"))[0]

    rh_pdf = list(pasta.glob("*relatorio_rh*.pdf"))[0]

    texto_freq = extrair_texto_pdf(freq_pdf)

    texto_rh = extrair_texto_pdf(rh_pdf)

    df_freq = parse_frequencia(texto_freq)

    df_rh = parse_rh(texto_rh)

    print("Linhas secretaria:", len(df_freq))

    print("Linhas RH:", len(df_rh))

    resultado = comparar(df_freq, df_rh)

    arquivo_saida = "comparacao_frequencia.xlsx"

    resultado.to_excel(arquivo_saida, index=False)

    colorir_excel(arquivo_saida)

    print("\nConferência finalizada!\n")

    print(resultado)


if __name__ == "__main__":
    main()